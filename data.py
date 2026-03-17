import requests
import json
import re
import os
import openpyxl
from bs4 import BeautifulSoup
import urllib3

# Suppress SSL security warnings when scraping websites with outdated certificates
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ================= CONFIGURATION =================
SERPER_API_KEY = "INSERT_YOUR_API_KEY_HERE"
MAX_ROWS_PER_SHEET = None  # Set to None to process all rows, or an integer for testing

NOME_FILE_INPUT = "rui_intermediari_A_B_E_collaborazioni.xlsx"
OUTPUT_FILE = "rui_intermediari_arricchito.xlsx"
NOMI_FOGLI = ["Tutti", "A", "B", "E"]

# Scraping parameters
regex_email = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
BLACKLIST = [
    'paginegialle.it', 'paginebianche.it', 'registroimprese.it', 'informazione-aziende.it', 
    'ufficiocamerale.it', 'icribis.com', 'facebook.com', 'instagram.com', 'linkedin.com', 
    'google.com', 'yelp.com', 'tripadvisor.com', 'kompass.com', 'prontoimprese.it'
]
# ==================================================

# Cache to minimize API calls for recurring RUI entries
cache_risultati = {}

def pulisci_nome(nome):
    """Removes corporate designations to optimize search accuracy."""
    termini_da_rimuovere = [
        "SOCIETA' A RESPONSABILITA' LIMITATA", "SOCIETA' PER AZIONI", 
        "SOCIETA' IN NOME COLLETTIVO", "SOCIETA' COOPERATIVA",
        "S.R.L.", "SRL", "S.N.C.", "SNC", "S.P.A.", "SPA", "S.A.S.", "SAS", "S.R.L.S."
    ]
    nome_pulito = nome.upper()
    for termine in termini_da_rimuovere:
        nome_pulito = nome_pulito.replace(termine, "")
    return " ".join(nome_pulito.split())

def cerca_su_google_maps(query):
    """Queries Google Maps via Serper API to fetch website and phone number data."""
    url = "https://google.serper.dev/places"
    payload = json.dumps({"q": query, "gl": "it", "hl": "it"})
    headers = {'X-API-KEY': SERPER_API_KEY, 'Content-Type': 'application/json'}
    try:
        response = requests.request("POST", url, headers=headers, data=payload)
        response.raise_for_status()
        dati = response.json()
        if 'places' in dati and len(dati['places']) > 0:
            primo_risultato = dati['places'][0]
            return {
                'sito': primo_risultato.get('website', None),
                'telefono': primo_risultato.get('phoneNumber', "non trovato")
            }
        return {'sito': None, 'telefono': "non trovato"}
    except Exception as e:
        return {'sito': None, 'telefono': "non trovato"}

def cerca_linkedin_serper(query):
    """Queries Google via Serper API to locate LinkedIn profiles."""
    url = "https://google.serper.dev/search"
    
    # Append 'site:linkedin.com' to restrict search results strictly to LinkedIn profiles
    payload = json.dumps({"q": f"{query} site:linkedin.com", "gl": "it", "hl": "it", "num": 3})
    headers = {'X-API-KEY': SERPER_API_KEY, 'Content-Type': 'application/json'}
    try:
        response = requests.request("POST", url, headers=headers, data=payload)
        dati = response.json()
        if 'organic' in dati and len(dati['organic']) > 0:
            
            # Filter results to verify the URL belongs to LinkedIn
            for res in dati['organic']:
                if "linkedin.com" in res['link']:
                    return res['link']
        return "non trovato"
    except:
        return "non trovato"

def estrai_email_da_sito(url):
    """Navigates to the provided URL to extract email addresses using regular expressions."""
    if any(dominio in url.lower() for dominio in BLACKLIST):
        print(f"      [!] Site skipped (Blacklisted): {url}")
        return "non trovato"
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        }
        
        # Enforce a 5-second timeout to prevent execution stalls on unresponsive websites
        response = requests.get(url, headers=headers, timeout=5, verify=False)
        response.raise_for_status() 
        testo_pagina = BeautifulSoup(response.text, 'html.parser').get_text()
        emails = list(set(re.findall(regex_email, testo_pagina)))
        emails_pulite = [e for e in emails if not e.endswith(('.png', '.jpg', '.gif', 'wixpress.com')) and "sentry" not in e]
        return emails_pulite[0].lower() if emails_pulite else "non trovato"
    except Exception as e:
        print(f"      [!] Navigation failed for site: {url}")
        return "non trovato"

def processa_riga(rui, nome, citta):
    """Processes a single spreadsheet row to fetch contact and LinkedIn data."""
    rui = str(rui).strip() if rui else ""
    nome = str(nome).strip() if nome else ""
    citta = str(citta).strip() if citta else ""
    
    if rui and rui in cache_risultati:
        print(f"[{rui}] -> Data retrieved from cache.")
        return cache_risultati[rui]
        
    if not nome or nome.lower() == 'none':
        return {"email": "non trovato", "telefono": "non trovato", "linkedin": "non trovato"}

    nome_pulito = pulisci_nome(nome)
    
    print(f"\n[{rui}] Agency: {nome_pulito} ({citta})")
    
    # 1. Google Maps query
    query_maps = f"{nome_pulito} assicurazioni {citta}"
    dati_maps = cerca_su_google_maps(query_maps)
    telefono = dati_maps['telefono']
    sito = dati_maps['sito']
    
    # 2. Email extraction via website scraping
    email = "non trovato"
    if sito and sito != "None":
        print(f"      -> Website identified: {sito} (Initiating scraping...)")
        email = estrai_email_da_sito(sito)
    else:
        print("      -> No website identified via Google Maps.")

    # 3. Targeted LinkedIn query
    query_linkedin = f"{nome_pulito} {citta} assicurazioni"
    linkedin = cerca_linkedin_serper(query_linkedin)
    
    print(f"      [RESULTS] Email: {email} | Phone: {telefono} | LinkedIn: {linkedin}")
    
    risultato = {"email": email, "telefono": telefono, "linkedin": linkedin}
    if rui: 
        cache_risultati[rui] = risultato
    return risultato

def main():
    if not os.path.exists(NOME_FILE_INPUT):
        print(f"ERROR: The input file '{NOME_FILE_INPUT}' was not found.")
        return

    print("Loading the original Excel file into memory...")
    wb = openpyxl.load_workbook(NOME_FILE_INPUT)

    for nome_foglio in NOMI_FOGLI:
        if nome_foglio not in wb.sheetnames:
            continue
            
        print(f"\n========================================")
        print(f"---> Processing sheet: {nome_foglio}")
        print(f"========================================")
        ws = wb[nome_foglio]
        
        header_row_idx = None
        for r in range(1, 10): 
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).strip().lower() == "nome intermediario":
                    header_row_idx = r
                    break
            if header_row_idx:
                break
                
        if not header_row_idx:
            continue
            
        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row_idx, column=c).value
            if val:
                col_map[str(val).strip().lower()] = c
                
        if "linkedin" not in col_map:
            nuova_colonna_idx = ws.max_column + 1
            ws.cell(row=header_row_idx, column=nuova_colonna_idx).value = "linkedin"
            col_map["linkedin"] = nuova_colonna_idx
            
        righe_elaborate = 0
        for r in range(header_row_idx + 1, ws.max_row + 1):
            if MAX_ROWS_PER_SHEET and righe_elaborate >= MAX_ROWS_PER_SHEET:
                break
                
            idx_rui = col_map.get("numero iscrizione rui")
            idx_nome = col_map.get("nome intermediario")
            idx_citta = col_map.get("città")
            
            if not idx_nome: continue
            
            rui = ws.cell(row=r, column=idx_rui).value if idx_rui else ""
            nome = ws.cell(row=r, column=idx_nome).value
            citta = ws.cell(row=r, column=idx_citta).value if idx_citta else ""
            
            if not nome: continue
            
            contatti = processa_riga(rui, nome, citta)
            
            if "e-mail" in col_map:
                ws.cell(row=r, column=col_map["e-mail"]).value = contatti["email"]
                
            if "telefono" in col_map:
                ws.cell(row=r, column=col_map["telefono"]).value = contatti["telefono"]
                
            if "linkedin" in col_map:
                ws.cell(row=r, column=col_map["linkedin"]).value = contatti["linkedin"]
                
            righe_elaborate += 1

    print(f"\nSaving file... Please do not terminate the script.")
    wb.save(OUTPUT_FILE)
    print(f"Process complete. The updated file has been successfully saved as: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()