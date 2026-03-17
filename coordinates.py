import requests
import json
import os
import openpyxl

# Configuration parameters
SERPER_API_KEY = "INSERT_YOUR_API_KEY_HERE"
MAX_ROWS_PER_SHEET = None  # Set to None to process all rows, or an integer for testing

NOME_FILE_INPUT = "rui_intermediari_A_B_E_collaborazioni.xlsx"
OUTPUT_FILE = "rui_intermediari_coordinate.xlsx"
NOMI_FOGLI = ["Tutti", "A", "B", "E"]

# Cache to minimize API calls for recurring RUI entries
cache_coordinate = {}

def pulisci_nome(nome):
    """Removes corporate designations to optimize Google search accuracy."""
    termini_da_rimuovere = [
        "SOCIETA' A RESPONSABILITA' LIMITATA", "SOCIETA' PER AZIONI", 
        "SOCIETA' IN NOME COLLETTIVO", "SOCIETA' COOPERATIVA",
        "S.R.L.", "SRL", "S.N.C.", "SNC", "S.P.A.", "SPA", "S.A.S.", "SAS", "S.R.L.S."
    ]
    nome_pulito = str(nome).upper()
    for termine in termini_da_rimuovere:
        nome_pulito = nome_pulito.replace(termine, "")
    return " ".join(nome_pulito.split())

def ottieni_coordinate_maps(query):
    """Queries Google Maps via Serper API to fetch latitude and longitude."""
    url = "https://google.serper.dev/places"
    
    # Strip newlines, tabs, and redundant spaces from the query
    query_pulita = " ".join(query.split())
    
    payload = {"q": query_pulita, "gl": "it", "hl": "it"}
    headers = {
        'X-API-KEY': SERPER_API_KEY,
        'Content-Type': 'application/json'
    }
    
    try:
        # Enforce a 10-second timeout to handle connection stalls
        response = requests.post(url, headers=headers, json=payload, timeout=10)
        
        # Validate response and check API credit limits
        if response.status_code != 200:
            if "Not enough credits" in response.text:
                print("\n🛑 [CRITICAL ERROR] Serper API credits exhausted. Please replace the API key.")
                exit() # Terminate execution if API credits are depleted
            print(f"      [!] Server error details: {response.text}")
            return "errore", "errore"
            
        dati = response.json()
        
        # Extract data from the 'places' array
        if 'places' in dati and len(dati['places']) > 0:
            primo_risultato = dati['places'][0]
            lat = primo_risultato.get('latitude', "non trovato")
            lng = primo_risultato.get('longitude', "non trovato")
            return str(lat), str(lng)
            
        return "non trovato", "non trovato"
        
    except requests.exceptions.Timeout:
        # Handle request timeout exceptions
        print("      [!] Request timed out (10s limit exceeded).")
        return "errore", "errore"
    except Exception as e:
        print(f"      [!] Exception encountered: {e}")
        return "errore", "errore"

def processa_riga_coordinate(rui, nome, citta):
    rui = str(rui).strip() if rui else ""
    nome = str(nome).strip() if nome else ""
    citta = str(citta).strip() if citta else ""
    
    if rui and rui in cache_coordinate:
        print(f"[{rui}] -> Coordinates retrieved from cache.")
        return cache_coordinate[rui]
        
    if not nome or nome.lower() == 'none':
        return {"lat": "non trovato", "lng": "non trovato"}

    nome_pulito = pulisci_nome(nome)
    
    query_maps = f"{nome_pulito} assicurazioni {citta}"
    print(f"\n[{rui}] Querying: {query_maps}")
    
    lat, lng = ottieni_coordinate_maps(query_maps)
    
    print(f"      -> Result: Lat: {lat} | Lng: {lng}")
    
    risultato = {"lat": lat, "lng": lng}
    if rui: 
        cache_coordinate[rui] = risultato
    return risultato

def main():
    if not os.path.exists(NOME_FILE_INPUT):
        print(f"ERROR: The input file '{NOME_FILE_INPUT}' was not found.")
        return

    print("Loading the original Excel file into memory...")
    wb = openpyxl.load_workbook(NOME_FILE_INPUT)

    for nome_foglio in NOMI_FOGLI:
        if nome_foglio not in wb.sheetnames:
            print(f"Warning: Sheet '{nome_foglio}' not found.")
            continue
            
        print(f"\n========================================")
        print(f"---> Extracting coordinates for sheet: {nome_foglio}")
        print(f"========================================")
        ws = wb[nome_foglio]
        
        header_row_idx = None
        for r in range(1, 10): 
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).strip().lower() == "nome intermediario":
                    header_row_idx = r
                    break
            if header_row_idx:
                break
                
        if not header_row_idx:
            continue
            
        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row_idx, column=c).value
            if val:
                col_map[str(val).strip().lower()] = c
                
        if "latitudine" not in col_map:
            col_lat = ws.max_column + 1
            ws.cell(row=header_row_idx, column=col_lat).value = "latitudine"
            col_map["latitudine"] = col_lat
            
        if "longitudine" not in col_map:
            col_lng = ws.max_column + 1
            ws.cell(row=header_row_idx, column=col_lng).value = "longitudine"
            col_map["longitudine"] = col_lng
            
        righe_elaborate = 0
        for r in range(header_row_idx + 1, ws.max_row + 1):
            if MAX_ROWS_PER_SHEET and righe_elaborate >= MAX_ROWS_PER_SHEET:
                break
                
            idx_rui = col_map.get("numero iscrizione rui")
            idx_nome = col_map.get("nome intermediario")
            idx_citta = col_map.get("città")
            
            if not idx_nome: continue
            
            rui = ws.cell(row=r, column=idx_rui).value if idx_rui else ""
            nome = ws.cell(row=r, column=idx_nome).value
            citta = ws.cell(row=r, column=idx_citta).value if idx_citta else ""
            
            if not nome: continue
            
            coordinate = processa_riga_coordinate(rui, nome, citta)
            
            if "latitudine" in col_map:
                ws.cell(row=r, column=col_map["latitudine"]).value = coordinate["lat"]
                
            if "longitudine" in col_map:
                ws.cell(row=r, column=col_map["longitudine"]).value = coordinate["lng"]
                
            righe_elaborate += 1

    print("\nSaving file... Please wait.")
    wb.save(OUTPUT_FILE)
    print(f"Process complete. The output file has been saved as: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()